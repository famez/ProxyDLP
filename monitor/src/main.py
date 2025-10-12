import re
import os
import pymupdf
from openpyxl import load_workbook
from docx import Document
from PIL import Image
import pytesseract
import io
import zipfile
from sentence_transformers import SentenceTransformer, util
#import spacy
from pymongo import MongoClient, ReturnDocument, ASCENDING
from bson.objectid import ObjectId
import numpy as np
import grpc
from concurrent import futures
import monitor_pb2
import monitor_pb2_grpc
from grpc_health.v1 import health, health_pb2, health_pb2_grpc
import faiss
import yara
from readerwriterlock import rwlock
from collections import Counter
from datetime import datetime, timezone
import logging
import logging.handlers
import json
from syslog_rfc5424_formatter import RFC5424Formatter
import smtplib
from email.mime.text import MIMEText
import json

import nltk
from nltk.corpus import stopwords
from sklearn.feature_extraction.text import TfidfVectorizer
import numpy as np

import hashlib


import threading
import time


INDEX_PATH = '/var/faiss/faiss_index.index'

background_executor = futures.ThreadPoolExecutor(max_workers=15)

db_client = MongoClient(os.getenv("MONGO_URI"))
events_collection = db_client["ProxyDLP"]["events"]
regex_collection = db_client["ProxyDLP"]["regex_rules"]
topics_collection = db_client["ProxyDLP"]["topic_rules"]
counter_collection = db_client["ProxyDLP"]["faiss_id_counters"]
yara_rules_collection = db_client["ProxyDLP"]["yara_rules"]

alert_destinations_collection = db_client["ProxyDLP"]["alert-destinations"]
alert_rules_collection = db_client["ProxyDLP"]["alert-rules"]
alert_locallogs_collection = db_client["ProxyDLP"]["alert-logs"]

nltk.download('stopwords')

#nlp = spacy.load("en_core_web_sm")

# Global shared resources and RWLocks
regex_rules = {}
regex_rw_lock = rwlock.RWLockFair()

faiss_index = None
faiss_rw_lock = rwlock.RWLockFair()

yara_rules_compiled = None
yara_rw_lock = rwlock.RWLockFair()

embeddings_model = SentenceTransformer('all-MiniLM-L6-v2')


def load_regex_rules():
    global regex_rules
    with regex_rw_lock.gen_wlock():
        rules = {}
        for doc in regex_collection.find():
            for regex_name, regex_value in doc.items():
                if regex_name != "_id":
                    rules[regex_name] = re.compile(regex_value)
        regex_rules = rules

def load_yara_rules():
    global yara_rules_compiled
    rule_sources = {}
    for doc in yara_rules_collection.find():
        try:
            rule_sources[doc['name']] = doc['content']
        except KeyError:
            print(f"Invalid YARA rule doc: {doc}")

    yara_rules_compiled = yara.compile(sources=rule_sources)
    with yara_rw_lock.gen_wlock():
        yara_rules_compiled.save("/tmp/yara_rules.yara")

def load_faiss_index():
    global faiss_index
    with faiss_rw_lock.gen_wlock():
        if os.path.exists(INDEX_PATH):
            faiss_index = faiss.read_index(INDEX_PATH)
        else:
            dim = 384  # Embedding size
            flat = faiss.IndexFlatIP(dim)
            faiss_index = faiss.IndexIDMap2(flat)


def match_regex(text):
    matches = {}
    with regex_rw_lock.gen_rlock():
        for name, pattern in regex_rules.items():
            for match in pattern.finditer(text):
                matches[match.group()] = name
    return matches

def match_yara(text):
    leaks = []

    # Acquire read lock
    with yara_rw_lock.gen_rlock():
        compiled_copy = yara.load("/tmp/yara_rules.yara")

    # Now perform the match
    matches = compiled_copy.match(data=text)
    for match in matches:
        matched_strings = []
        for string_match in match.strings:
            identifier = string_match.identifier
            for instance in string_match.instances:
                matched_strings.append({
                    "offset": instance.offset,
                    "identifier": identifier,
                    "data": instance.matched_data.decode(errors="ignore")
                })
        leaks.append({
            "name": match.rule,
            "matched_strings": matched_strings,
            "tags": match.tags,
            "meta": match.meta
        })
    return leaks


def search_faiss(embedding_vector):
    with faiss_rw_lock.gen_rlock():
        scores, indices = faiss_index.search(embedding_vector.reshape(1, -1), 10)
    return scores, indices

def reload_all_rules():
    load_regex_rules()
    load_yara_rules()
    load_faiss_index()

def get_next_faiss_id():
    counter = counter_collection.find_one_and_update(
        {"_id": "faiss_id_counter"},
        {"$inc": {"last_id": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER
    )
    return counter["last_id"]

def extract_rule_identifiers(rule_str):
    # Match rule names using regex: rule <identifier>
    pattern = r'\brule\s+(\w+)\s*{'
    return re.findall(pattern, rule_str)

def validate_yara_rule_string(rule_str):
    try:
        yara.compile(source=rule_str)
        rule_identifiers = extract_rule_identifiers(rule_str)
        print("YARA rule is valid.")
        return True, rule_identifiers
    except yara.SyntaxError as e:
        print(f"Syntax error: {e}")
        return False, []
    except Exception as e:
        print(f"Error: {e}")
        return False, []



def chunk_text(text, chunk_size=500, overlap=50):
    """Split text into chunks of `chunk_size` with optional `overlap`."""
    words = text.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        chunk = ' '.join(words[i:i + chunk_size])
        chunks.append(chunk)
    return chunks


def analyze_topic_leak(text):

    leaked_topics = []
    chunk_embeddings = obtain_embeddings_from_text(text)
    
    # Let's check if the is similarity with the faiss index
    embedding_vectors = np.array([emb['embedding'] for emb in chunk_embeddings], dtype='float32')
    for embedding in embedding_vectors:
        #faiss_index.hnsw.efSearch = 16  # Query time accuracy/speed tradeoff, default is 16
        scores, indices = search_faiss(embedding)  # This will use the global faiss_index
        for score, idx in zip(scores[0], indices[0]):
            if score >= 0.3:        #Cosine similarity threshold higher or equals to 0.3
                print(f"Found similar embedding with score {score} at index {idx}")

                doc = topics_collection.find_one(
                    {"faiss_indexes": int(idx)},
                    {"faiss_indexes.$": 1, "name": 1}  # Only project the matched index with the chunk
                )

                if doc:
                    print("Matched Document ID:", doc["_id"])
                    leaked_topics.append({"name": doc['name'], "faiss_id": int(idx), "score": float(score)})

                else:
                    print("No matching document found.")

    return leaked_topics

    
def obtain_embeddings_from_text(text):
    chunks = chunk_text(text)
    embeddings = embeddings_model.encode(chunks, normalize_embeddings=True)

    linked_data = [{"chunk": chunk, "embedding": embedding.tolist()} for chunk, embedding in zip(chunks, embeddings)]

    return linked_data



def analyze_text(text):
    leak = {}
    leak['regex'] = analyze_text_regex(text)
    leak['topic'] = analyze_topic_leak(text)
    leak['yara'] = analyze_text_yara(text)

    return leak

def analyze_text_regex(text):

    return match_regex(text)
        


def analyze_text_yara(text):

    return match_yara(text)


def decode_file(filepath, content_type):

    text = ""
    if content_type == "application/pdf":
        with pymupdf.open(filepath) as doc:  # open document
            text = chr(12).join([page.get_text() for page in doc])

            #Extract images and apply OCR
            for page_num in range(len(doc)):
                page = doc[page_num]
                image_list = page.get_images(full=True)
                print(f"[+] Found {len(image_list)} images on page {page_num}")

                for img_index, img in enumerate(image_list):
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    image_bytes = base_image["image"]
                    image = Image.open(io.BytesIO(image_bytes))
                    text += pytesseract.image_to_string(image)
                                    
    elif content_type == "application/vnd.ms-excel" or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        workbook = load_workbook(filename=filepath)
        text_data = []
    
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = ' '.join([str(cell) for cell in row if cell is not None])
                text_data.append(row_text)
        
        text = '\n'.join(text_data)

    elif content_type == "application/msword" or content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        doc = Document(filepath)
        text = [para.text for para in doc.paragraphs if para.text.strip()]
        text = '\n'.join(text)

        extract_images_from_docx(filepath)


    elif content_type == "image/jpeg" or content_type == "image/png" or content_type == "image/gif" or content_type == "image/bmp" or content_type == "image/webp" or content_type == "image/svg+xml" or content_type == "image/tiff" or content_type == "image/vnd.microsoft.icon":
        
        image = Image.open(filepath)
        text = pytesseract.image_to_string(image)

    elif content_type == "text/plain":
        with open(filepath, 'r') as file:
            text = file.read()
    
    return text


def sha256_hash_file(filename):
    sha256_hash = hashlib.sha256()
    with open(filename, "rb") as f:
        # Read and update hash in chunks (good for large files)
        for chunk in iter(lambda: f.read(4096), b""):
            sha256_hash.update(chunk)
    return sha256_hash.hexdigest()


def analyze_file(filepath, content_type):

    #Calculate sha256
    file_hash = sha256_hash_file(filepath)

    print(f"File hash: {file_hash}")

    text = decode_file(filepath, content_type)
    return text, analyze_text(text), file_hash



def extract_images_from_docx(docx_path, output_folder="extracted_images"):
    with zipfile.ZipFile(docx_path, 'r') as docx_zip:
        # Create output folder if it doesn't exist
        os.makedirs(output_folder, exist_ok=True)

        # Loop through files in the ZIP and extract images
        for file in docx_zip.namelist():
            if file.startswith("word/media/"):
                filename = os.path.basename(file)
                if filename:  # skip folders
                    target_path = os.path.join(output_folder, filename)
                    with open(target_path, "wb") as img_file:
                        img_file.write(docx_zip.read(file))
                    print(f"Saved image: {target_path}")


def on_event_added(event_id):
    try:
        print(f"Started on_event_added for Event ID: {event_id}")
        event = events_collection.find_one({'_id': ObjectId(event_id)})

        print(f"Event obtained")
    
        leak = {}
        result = {}

        if event['rational'] == "Conversation":
            #print(f"Conversation, analysing: {event['content']}")
            leak = analyze_text(event['content'])
            #print(f"Done: {leak}")


            result = events_collection.update_one(
                {"_id": ObjectId(event_id)},
                {"$set": {"leak": leak}}
            )

        elif event['rational'] == "Attached file":
            text, leak, hash = analyze_file(event['filepath'], event['content_type'])

            result = events_collection.update_one(
                {"_id": ObjectId(event_id)},
                {"$set": {"leak": leak, "hash": hash}}
            )


        if result.modified_count > 0:
            print("Document updated successfully.")
        else:
            print("No changes made or document not found.")

        check_alerts(leak)
        

        print(f"Finished long task for Event ID: {event_id}")

    except Exception as e:
        # Handle the exception
        print(f"An error occurred: {e}")


def check_alerts(leak):

    # Count regex values
    regex_counts = Counter(leak["regex"].values())

    # Count topic names
    topic_counts = Counter(entry["name"] for entry in leak["topic"])

    # Count yara names
    yara_counts = Counter(entry["name"] for entry in leak["yara"])

    leak_count = {
        "regex": dict(regex_counts),
        "topic": dict(topic_counts),
        "yara": dict(yara_counts)
    }

    print(leak_count)

    results = list(alert_rules_collection.aggregate([
        {
            "$lookup": {
                "from": "topic_rules",
                "localField": "topic.rules",
                "foreignField": "_id",
                "as": "resolved_topic"
            }
        },
        {
            "$lookup": {
                "from": "yara_rules",
                "localField": "yara.rules",
                "foreignField": "_id",
                "as": "resolved_yara"
            }
        },
        {
            "$lookup": {
                "from": "regex_rules",
                "localField": "regex.rules",
                "foreignField": "_id",
                "as": "resolved_regex"
            }
        },
        {
            "$lookup": {
                "from": 'alert-destinations',
                "localField": 'destinations',
                "foreignField": '_id',
                "as": 'destinationResolved'
            }
        },
        {
            "$project": {
                "name": 1,
                "topic_names": "$resolved_topic.name",
                "topic_count": "$topic.count",
                "yara_ids": {
                    "$reduce": {
                        "input": "$resolved_yara",
                        "initialValue": [],
                        "in": {
                        "$concatArrays": ["$$value", "$$this.identifiers"]
                        }
                    }
                },
                "yara_count": "$yara.count",
                "regex_names": {
                    "$reduce": {
                        "input": "$resolved_regex",
                        "initialValue": [],
                        "in": {
                        "$concatArrays": [
                            "$$value",
                            {
                            "$map": {
                                "input": {
                                "$filter": {
                                    "input": { "$objectToArray": "$$this" },
                                    "as": "item",
                                    "cond": { "$ne": ["$$item.k", "_id"] }
                                }
                                },
                                "as": "item",
                                "in": "$$item.k"
                            }
                            }
                        ]
                        }
                    }
                },
                "regex_count": "$regex.count",
                "destinations": '$destinationResolved'
            }
        }
    ]))

    print (results)

    # Comparison function
    def rule_matches(alert_doc, leak_count):
        for name in alert_doc.get("topic_names", []):
            if leak_count["topic"].get(name, 0) < alert_doc.get("topic_count", 0):
                return False
        for name in alert_doc.get("yara_ids", []):
            if leak_count["yara"].get(name, 0) < alert_doc.get("yara_count", 0):
                return False
        for name in alert_doc.get("regex_names", []):
            if leak_count["regex"].get(name, 0) < alert_doc.get("regex_count", 0):
                return False
        return True
    

    # Filter matching alerts
    matching_alerts = [doc for doc in results if rule_matches(doc, leak_count)]

    # Send matching alert rules to the configured destinations
    for alert in matching_alerts:
        print(f"Matching alert rule: {alert['name']}")
        for destination in alert["destinations"]:
            if(destination['type'] == "local_logs"):
                rotation_limit = destination.get("rotationLimit", 500)
                send_alert_to_local_logs(alert, leak, rotation_limit)
            elif(destination['type'] == "syslog"):
                send_alert_to_syslog(alert, destination, leak)
            elif(destination['type'] == "email"):
                send_alert_to_email(alert, destination, leak)


def send_alert_to_local_logs(alert, leak, rotation_limit):
    
    alert_locallogs_collection.insert_one(
        {
            "timestamp": datetime.now(timezone.utc),
            "alert_rule": alert['name'],
            "leak": leak
        }
    )


    # Count current number of logs
    total_logs = alert_locallogs_collection.count_documents({})

    # Remove oldest logs if over limit
    if total_logs > rotation_limit:
        to_delete = total_logs - rotation_limit

        oldest_logs = alert_locallogs_collection.find({}, {"_id": 1}).sort("timestamp", ASCENDING).limit(to_delete)
        ids_to_delete = [doc["_id"] for doc in oldest_logs]

        alert_locallogs_collection.delete_many({"_id": {"$in": ids_to_delete}})


def send_alert_to_syslog(alert, destination, leak):
    # Create a syslog handler
    print("Send to syslog...")
    #print(f"Alert: {alert}")
    logger = logging.getLogger('ProxyDLP')
    logger.setLevel(logging.INFO)
    syslog_handler = logging.handlers.SysLogHandler(address=(destination['syslogHost'], int(destination['syslogPort'])), facility=logging.handlers.SysLogHandler.LOG_USER)
    
    formatter = RFC5424Formatter()
    syslog_handler.setFormatter(formatter)
    logger.addHandler(syslog_handler)

    log_data = {
        "alert_rule": alert['name'],
        "leak": leak
    }

    logger.info(json.dumps(log_data))


def send_alert_to_email(alert, destination, leak):
    smtp_host = destination.get("smtpHost")
    smtp_port = int(destination.get("smtpPort", 25))
    smtp_user = destination.get("email")
    smtp_pass = destination.get("emailPassword")
    recipient = destination.get("recipientEmail")

    # Format message
    subject = f"[Alert] {alert['name']}"

    # Generate HTML body similar to the EJS layout
    body = f"""
    <html>
    <head>
      <style>
        body {{
          font-family: sans-serif;
          background-color: #fdfdfd;
          padding: 20px;
        }}
        .container {{
          max-width: 800px;
          margin: 0 auto;
          background: #fff;
          padding: 20px;
          border-radius: 8px;
          box-shadow: 0 2px 6px rgba(0,0,0,0.1);
        }}
        h2 {{
          color: #b83280;
        }}
        table {{
          width: 100%;
          border-collapse: collapse;
        }}
        th, td {{
          border: 1px solid #e2e8f0;
          padding: 8px;
          text-align: left;
        }}
        thead {{
          background-color: #fed7e2;
          color: #9d174d;
        }}
      </style>
    </head>
    <body>
      <div class="container">
        <p>ProxyDLP alert</p>
        <table>
          <thead>
            <tr>
              <th>Time</th>
              <th>Alert Rule</th>
              <th>Matched Yara</th>
              <th>Regex Matches</th>
              <th>Matched Topics</th>
            </tr>
          </thead>
          <tbody>
            <tr>
              <td>{leak.get('time', 'N/A')}</td>
              <td>{alert.get('name')}</td>
              <td>{leak.get('yara', 'N/A')}</td>
              <td>{leak.get('regex', 'N/A')}</td>
              <td>{leak.get('topic', 'N/A')}</td>
            </tr>
          </tbody>
        </table>
      </div>
    </body>
    </html>
    """

    # Prepare HTML email
    msg = MIMEText(body, 'html')
    msg["Subject"] = subject
    msg["From"] = smtp_user
    msg["To"] = recipient

    # Send email
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.set_debuglevel(1)
        try:
            server.login(smtp_user, smtp_pass)
        except smtplib.SMTPException:
            pass  # Allow no-auth servers like MailHog
        server.sendmail(smtp_user, [recipient], msg.as_string())



def on_topic_rule_added(topic_rule_id):
    try:

        print(f"Started on_topic_rule_added for Topic Rule ID: {topic_rule_id}")
        topic_rule = topics_collection.find_one({'_id': ObjectId(topic_rule_id)})

        print(f"Topic Rule obtained: {topic_rule}")

        chunks = chunk_text(topic_rule['pattern'])
        embeddings = embeddings_model.encode(chunks, normalize_embeddings=True)
        faiss_indexes = [get_next_faiss_id() for chunk in chunks]

        # Prepare data
        ids = np.array(faiss_indexes, dtype='int64')

        # Add embeddings to FAISS index with write lock
        with faiss_rw_lock.gen_wlock():
            faiss_index.add_with_ids(embeddings, ids)
            # Save the FAISS index to disk
            faiss.write_index(faiss_index, INDEX_PATH)

        topics_collection.update_one(
            {'_id': ObjectId(topic_rule_id)},  # Filter to find the document
            {'$set': {'faiss_indexes': faiss_indexes}}  # Field to add or update
        )

        #print(f"Topic Rule encoded: {encoded}")

    except Exception as e:
        # Handle the exception
        print(f"An error occurred: {e}")


def remove_topic_rule(topic_rule_id, delete_only_indexes=False):
    print(f"Received Topic Rule ID: {topic_rule_id}")

    topic_rule = topics_collection.find_one({'_id': ObjectId(topic_rule_id)})

    try:

        ids = np.ascontiguousarray(np.array(topic_rule['faiss_indexes'], dtype=np.int64))
        selector = faiss.IDSelectorBatch(ids)
        with faiss_rw_lock.gen_wlock():
            n_removed = faiss_index.remove_ids(selector)
            print(f"Removed {n_removed} vectors from FAISS index")
            # Save the FAISS index to disk
            faiss.write_index(faiss_index, INDEX_PATH)

        if delete_only_indexes:
            topics_collection.update_one(
                {'_id': ObjectId(topic_rule_id)},
                {'$unset': {'faiss_indexes': ""}}
            )
        else:
            print(f"Deleting document")
            topics_collection.delete_one({'_id': ObjectId(topic_rule_id)})

    except Exception as e:
        print(f"Exception: {e}")

    
    

class MonitorServicer(monitor_pb2_grpc.MonitorServicer):

    def EventAdded(self, request, context):
        print(f"Received Event ID: {request.id}")
        background_executor.submit(on_event_added, request.id)
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)
    
    def TopicRuleAdded(self, request, context):
        print(f"Received Topic Rule ID: {request.id}")
        on_topic_rule_added(request.id)
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)
    
    def TopicRuleRemoved(self, request, context):
        remove_topic_rule(request.id, delete_only_indexes=False)
        
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)
    
    def TopicRuleEdited(self, request, context):
        remove_topic_rule(request.id, delete_only_indexes=True)
        on_topic_rule_added(request.id)
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)
    
    #We have this callback to check if the Yara rule is valid before saving it to the database
    def YaraRuleAdded(self, yara_rule, context):
        print(f"Received Yara rule name: {yara_rule.name}")

        is_valid, rule_identifiers = validate_yara_rule_string(yara_rule.content)
        if not is_valid:
            print(f"Invalid Yara rule: {yara_rule.name}")
            return monitor_pb2.MonitorReply(result=1)
        
        try:
            # Save the Yara rule to the database
            yara_rules_collection.insert_one({
                "name": yara_rule.name,
                "content": yara_rule.content,
                "identifiers": rule_identifiers
            })

            load_yara_rules()  # Reload Yara rules after adding a new one
        except Exception as e:
            print(f"Error saving Yara rule: {e}")
            return monitor_pb2.MonitorReply(result=2)
        
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)
    
    def YaraRuleEdited(self, yara_rule_edit_request, context):
        print(f"Received Yara rule name: {yara_rule_edit_request.rule.name}")
        print(f"Received Yara rule id: {yara_rule_edit_request.id.id}")

        is_valid, rule_identifiers = validate_yara_rule_string(yara_rule_edit_request.rule.content)
        if not is_valid:
            print(f"Invalid Yara rule: {yara_rule_edit_request.rule.name}")
            return monitor_pb2.MonitorReply(result=1)
        
        try:
            # Save the Yara rule to the database
            yara_rules_collection.update_one( 
                {'_id': ObjectId(yara_rule_edit_request.id.id)},  # Filter to find the document
                {'$set': 
                    {  
                        "name": yara_rule_edit_request.rule.name,
                        "content": yara_rule_edit_request.rule.content,
                        "identifiers": rule_identifiers
                    }
                }  # Field to add or update
            )

            load_yara_rules()  # Reload Yara rules after adding a new one
        except Exception as e:
            print(f"Error saving Yara rule: {e}")
            return monitor_pb2.MonitorReply(result=2)
        
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)

    def YaraRuleDeleted(self, yara_rule, context):
        load_yara_rules()  # Reload Yara rules after removing it
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)

    def RegexRuleAdded(self, regex_rule, context):
        load_regex_rules()
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)

    def RegexRuleRemoved(self, regex_rule, context):
        load_regex_rules()
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)

    def RegexRuleEdited(self, regex_rule, context):
        load_regex_rules()
        return monitor_pb2.MonitorReply(result=0)       #Everything ok :)


def perform_tf_idf():

    # Create combined stopwords set from all needed languages
    all_stopwords = set()
    for lang in ['english', 'spanish', 'french']:
        all_stopwords.update(stopwords.words(lang))

    all_stopwords = list(all_stopwords)

    # Extract all text from events and keep their _id for mapping
    event_ids = []
    texts = []
    for event in events_collection.find({}, {"_id": 1, "content": 1, "filepath": 1, "content_type": 1, "rational": 1}):
        text = ""
        if event.get("rational") == "Conversation" and event.get("content"):
            text = event["content"]
        elif event.get("rational") == "Attached file" and event.get("filepath") and event.get("content_type"):
            try:
                text = decode_file(event["filepath"], event["content_type"])
                
            except Exception as e:
                print(f"Error decoding file {event['filepath']}: {e}")

        #Check that we are indeed appending a string
        if isinstance(text, str) and text:
            texts.append(text)
            event_ids.append(event["_id"])

    if not texts:
        print("No events to process for TF-IDF.")
        return

    # Initialize vectorizer with combined stop words
    vectorizer = TfidfVectorizer(stop_words=all_stopwords)
    tfidf_matrix = vectorizer.fit_transform(texts)
    feature_names = vectorizer.get_feature_names_out()

    for i, (event_id, text) in enumerate(zip(event_ids, texts)):
        scores = tfidf_matrix[i].toarray().flatten()
        top_n = 5
        top_indices = np.argsort(scores)[::-1][:top_n]
        top_words = []
        for idx in top_indices:
            if scores[idx] > 0:
                word_score = {"word": feature_names[idx], "score": float(scores[idx])}
                top_words.append(word_score)
                
        # Store top words in the database for this event
        events_collection.update_one(
            {"_id": event_id},
            {"$set": {"tfidf_top_words": top_words}}
        )

def run_tf_idf_periodically():
    """Run perform_tf_idf immediately and then every 2 hours."""
    while True:
        try:
            perform_tf_idf()
        except Exception as e:
            print(f"Error in perform_tf_idf: {e}")
        time.sleep(2 * 60 * 60)  # Sleep for 2 hours



def main():

    reload_all_rules()  # Load all rules at startup

    # Start TF-IDF background thread
    threading.Thread(target=run_tf_idf_periodically, daemon=True).start()


    # Initialize gRPC server
    server = grpc.server(futures.ThreadPoolExecutor(max_workers=10))
    monitor_pb2_grpc.add_MonitorServicer_to_server(MonitorServicer(), server)

    #For health check to ensure proper start up of the containers
    # Add health service
    health_servicer = health.HealthServicer()
    health_pb2_grpc.add_HealthServicer_to_server(health_servicer, server)
    health_servicer.set('', health_pb2.HealthCheckResponse.SERVING)

    server.add_insecure_port("[::]:50051")
    server.start()
    print("Server running on port 50051...")
    server.wait_for_termination() 

if __name__ == "__main__":
    main()