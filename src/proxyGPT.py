# modify_headers.py
from mitmproxy import http, ctx
from mitmproxy.http import Response
import json
import re
import base64
import os
from datetime import datetime, timezone
import pymupdf
from openpyxl import load_workbook
from docx import Document
from PIL import Image
import pytesseract
import io
import zipfile
from sentence_transformers import SentenceTransformer, util
import spacy
from pymongo import MongoClient
from datetime import datetime


db_client = MongoClient(os.getenv("MONGO_URI"))
collection = db_client["proxyGPT"]["events"]



nlp = spacy.load("en_core_web_sm")

allowed_domain = "gmail.com"      #Change by your org domain, such as contoso.com

email_regex = r'^[a-zA-Z0-9._%+-]+@' + allowed_domain + '$'

#Open regex config file
with open("/etc/proxyGPT/regex.json") as f:
    regex_list = json.load(f)

regex_list = [(name, re.compile(pattern)) for name, pattern in regex_list]

source_code_patterns = [
    r'\b(def|function|class)\b\s+\w+\s*\(',     # function or class
    r'\bimport\s+\w+',                          # import
    r'\b\w+\s*=\s*.+',                          # assignment
    r'[{};]',                                   # semicolons/braces
    r'(//|#|/\*)'                               # comments
]

#Open cosine similarity strings config file
with open("/etc/proxyGPT/cos_similarity.json") as f:
    target_labels = json.load(f)
    

model = SentenceTransformer('all-MiniLM-L6-v2')

target_label_embeddings = []

for i, label in enumerate(target_labels):
    target_label_embeddings.append(model.encode(target_labels[i], convert_to_tensor=True))


def request(flow: http.HTTPFlow) -> None:
    # Example: Add a custom header to requests to example.com
    #print("Requested URL:", flow.request.pretty_url)
    #ctx.log.info("test")
    
    if flow.request.method == "POST" and "auth.openai.com/api/accounts/authorize/continue" in flow.request.pretty_url:
        ctx.log.info("Performing authentication!")
        json_body = flow.request.json()

        if 'connection' in json_body:
            #Don't allow delegated authentication
            ctx.log.info("Blocking delegated authentication!")
            # Define the JSON structure you want to return
            response_data = {
                "continue_url": "https://chatgpt.com",
                "method": "GET",
            }

            # Return JSON response
            flow.response = Response.make(
                200,
                json.dumps(response_data).encode("utf-8"),  # Must be bytes
                {"Content-Type": "application/json"}
            )
            return
    
        if 'username' in json_body and json_body['username']['kind'] == "email":

            email = json_body['username']["value"]
            ctx.log.info(f"Using email {email}")

            #Check whether the email address belongs to the organization or not
            if not re.match(email_regex, email):
                ctx.log.info(f"Email address does not belong to an organization")
                response_data = {
                    "continue_url": "https://chatgpt.com",
                    "method": "GET",
                }

                # Return JSON response
                flow.response = Response.make(
                    200,
                    json.dumps(response_data).encode("utf-8"),  # Must be bytes
                    {"Content-Type": "application/json"}
                )

            else:
                ctx.log.info(f"Corporative user {email} logged in")

                #Register event into the database.
                event = {"timestamp": datetime.now(datetime.timezone.utc), "user": email, "rational": "Logged in", "detail" : ""}
                collection.insert_one(event)

                return

    if flow.request.method == "POST" and "chatgpt.com/backend-anon/conversation" in flow.request.pretty_url:
        # Return JSON response

        ctx.log.info(f"Anonymous conversations are not allowed")
        flow.response = Response.make(
            403,
            b"Blocked by proxy",  # Body
            {"Content-Type": "text/plain"}  # Headers
        )
        return        
    
    if flow.request.method == "POST" and (flow.request.pretty_url == "https://chatgpt.com/backend-api/conversation"
                                          or flow.request.pretty_url == "https://chatgpt.com/backend-api/f/conversation"):
        
        ctx.log.info(f"Authenticated conversation...")

        #Obtain JWT token to double check that the session is still authorized
        auth_header = flow.request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            
            jwt_token = auth_header[len("Bearer "):].strip()
            #ctx.log.info(f"JWT Token extracted: {jwt_token}")

            jwt_data = decode_jwt(jwt_token)

            if jwt_data:
                #ctx.log.info(f"JWT Header: {json.dumps(jwt_data['header'], indent=2)}")
                #ctx.log.info(f"JWT Payload: {json.dumps(jwt_data['payload'], indent=2)}")

                jwt_payload = jwt_data['payload']

                #Let's check only the email address from the JWT token, 
                #as the rest of fields are already validated by Chatgpt to 
                #perform the request (correctly signed, not expired, etc).

                if "https://api.openai.com/profile" in jwt_payload and 'email' in jwt_payload["https://api.openai.com/profile"]:

                    email = jwt_payload["https://api.openai.com/profile"]['email']

                    if re.match(email_regex, email):
                        ctx.log.info(f"Email address belongs to the organization")

                        #Get the text sent to the conversation

                        json_body = flow.request.json()
                        conversation_text = json_body["messages"][0]["content"]["parts"][0]

                        ctx.log.info(f"Conversation sent: {conversation_text}")

                        result = analyze_text(conversation_text)

                        ctx.log.info(f"Leaked data from conversation: {result}")

                        event = {"timestamp": datetime.now(timezone.utc), "user": email, "rational": "Conversation", "content": conversation_text,"leak" : result}

                        collection.insert_one(event)

                        return

            ctx.log.info("JWT token checks failed!")
            flow.response = Response.make(
                403,
                b"Blocked by proxy",  # Body
                {"Content-Type": "text/plain"}  # Headers
            )


    #File being uploaded to ChatGPT.
    if flow.request.method == "PUT" and "oaiusercontent.com/file" in flow.request.pretty_url:

        content = flow.request.content
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if content:
            # Generate a filename from URL or timestamp
            filename = f"{timestamp}"

            content_type = flow.request.headers.get("Content-Type", "unknown")
            if content_type == "application/pdf":
                filename += ".pdf"
            elif content_type == "application/vnd.ms-excel" or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                filename += ".xlsx"
            elif content_type == "application/msword" or content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                filename += ".docx"
            elif content_type == "image/jpeg":
                filename += ".jpg"
            elif content_type == "image/png":
                filename += ".png"
            elif content_type == "image/gif":
                filename += ".gif"
            elif content_type == "image/bmp":
                filename += ".bmp"
            elif content_type == "image/webp":
                filename += ".webp"
            elif content_type == "image/svg+xml":
                filename += ".svg"
            elif content_type == "image/tiff":
                filename += ".tiff"
            elif content_type == "image/vnd.microsoft.icon":
                filename += ".ico"

            filepath = os.path.join("uploads", filename)

            os.makedirs("uploads", exist_ok=True)

            with open(filepath, "wb") as f:
                f.write(content)

            ctx.log.info(f"Saved PUT upload to: {filepath}")

            ctx.log.info(f"Analysing file...")
            text, result = analyze_file(filepath, content_type)

            ctx.log.info(f"Leaked data from file: {result}")

            event = {"timestamp": datetime.now(timezone.utc), "rational": "Conversation", "content": text,"leak" : result}
            collection.insert_one(event)
          

def decode_jwt(token: str):
    parts = token.split(".")
    if len(parts) != 3:
        return None
    try:
        header = json.loads(base64.urlsafe_b64decode(pad_b64(parts[0])).decode())
        payload = json.loads(base64.urlsafe_b64decode(pad_b64(parts[1])).decode())
        return {"header": header, "payload": payload}
    except Exception as e:
        ctx.log.warn(f"JWT decoding error: {str(e)}")
        return None

def pad_b64(segment: str) -> str:
    return segment + '=' * (-len(segment) % 4)


def analyze_text_ner(text):
    doc = nlp(text)
    return " ".join(ent.label_ for ent in doc.ents)
        
    
def analyze_text_cosine_similarity(text):
    feature_embedding = model.encode(text, convert_to_tensor=True)
    for i, label_embedding in enumerate(target_label_embeddings):
        similarity = util.cos_sim(feature_embedding, label_embedding).item()
        if similarity > 0.40:
            return target_labels[i]
    return ""


def analyze_text(text):
    retVal = ""
    retVal += analyze_text_regex(text) + " "
    retVal += analyze_text_ner(text) + " "
    retVal += analyze_text_cosine_similarity(text) + " "
    return retVal

#Returns string of leak type, empty string otherwise.
def analyze_text_regex(text):

    for line in text.splitlines():
        for regex in regex_list:
            if re.match(regex[1], line):
                return regex[0]
        
    return ""

def decode_file(filepath, content_type):

    text = ""
    if content_type == "application/pdf":
        with pymupdf.open(filepath) as doc:  # open document
            text = chr(12).join([page.get_text() for page in doc])

            #Extract images and apply OCR
            for page_num in range(len(doc)):
                page = doc[page_num]
                image_list = page.get_images(full=True)
                print(f"[+] Found {len(image_list)} images on page {page_num}")

                for img_index, img in enumerate(image_list):
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    image_bytes = base_image["image"]
                    image = Image.open(io.BytesIO(image_bytes))
                    text += pytesseract.image_to_string(image)
                                    
    elif content_type == "application/vnd.ms-excel" or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        workbook = load_workbook(filename=filepath)
        text_data = []
    
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = ' '.join([str(cell) for cell in row if cell is not None])
                text_data.append(row_text)
        
        text = '\n'.join(text_data)

    elif content_type == "application/msword" or content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        doc = Document(filepath)
        text = [para.text for para in doc.paragraphs if para.text.strip()]
        text = '\n'.join(text)

        extract_images_from_docx(filepath)


    elif content_type == "image/jpeg" or content_type == "image/png" or content_type == "image/gif" or content_type == "image/bmp" or content_type == "image/webp" or content_type == "image/svg+xml" or content_type == "image/tiff" or content_type == "image/vnd.microsoft.icon":
        
        image = Image.open(filepath)
        text = pytesseract.image_to_string(image)


    
    return text


def analyze_file(filepath, content_type):
    text = decode_file(filepath, content_type)
    return text, analyze_text(text)



def extract_images_from_docx(docx_path, output_folder="extracted_images"):
    with zipfile.ZipFile(docx_path, 'r') as docx_zip:
        # Create output folder if it doesn't exist
        os.makedirs(output_folder, exist_ok=True)

        # Loop through files in the ZIP and extract images
        for file in docx_zip.namelist():
            if file.startswith("word/media/"):
                filename = os.path.basename(file)
                if filename:  # skip folders
                    target_path = os.path.join(output_folder, filename)
                    with open(target_path, "wb") as img_file:
                        img_file.write(docx_zip.read(file))
                    print(f"Saved image: {target_path}")
